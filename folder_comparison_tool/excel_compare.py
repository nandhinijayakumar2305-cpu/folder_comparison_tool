import pandas as pd
from openpyxl import load_workbook


def get_sheet_names(filepath):
    try:
        wb = load_workbook(filepath, read_only=True)
        return wb.sheetnames
    except Exception:
        return []


def compare_sheets(df1, df2, sheet_name):
    differences = []

    # Fill NaN properly
    df1 = df1.fillna("")
    df2 = df2.fillna("")

    # Row count check
    if len(df1) != len(df2):
        differences.append({
            "type": "Row Count Changed",
            "detail": f"{sheet_name}: {len(df1)} → {len(df2)} rows"
        })

    # Column comparison
    cols1 = set(df1.columns)
    cols2 = set(df2.columns)

    added_cols = cols2 - cols1
    removed_cols = cols1 - cols2

    for col in added_cols:
        differences.append({
            "type": "Column Added",
            "detail": f"{sheet_name}: '{col}' added"
        })

    for col in removed_cols:
        differences.append({
            "type": "Column Removed",
            "detail": f"{sheet_name}: '{col}' removed"
        })

    # Common columns (SORTED → important fix)
    common_cols = sorted(list(cols1 & cols2))

    # Align rows
    min_rows = min(len(df1), len(df2))

    df1_common = df1[common_cols].iloc[:min_rows].reset_index(drop=True)
    df2_common = df2[common_cols].iloc[:min_rows].reset_index(drop=True)

    # Convert to string safely
    df1_str = df1_common.astype(str)
    df2_str = df2_common.astype(str)

    # Vectorized diff (FAST)
    diff_positions = (df1_str != df2_str)

    changed_cells = diff_positions.stack()
    changed_cells = changed_cells[changed_cells]

    for (row_idx, col) in changed_cells.index:
        differences.append({
            "type": "Cell Changed",
            "detail": (
                f"{sheet_name} | Row {row_idx + 2} | Column '{col}' | "
                f"{df1_str.at[row_idx, col]} → {df2_str.at[row_idx, col]}"
            )
        })

    return differences


def compare_excel_files(file1, file2):
    result = {
        "sheets_added": [],
        "sheets_removed": [],
        "differences": []
    }

    try:
        sheets1 = set(get_sheet_names(file1))
        sheets2 = set(get_sheet_names(file2))

        result["sheets_added"] = list(sheets2 - sheets1)
        result["sheets_removed"] = list(sheets1 - sheets2)

        common_sheets = sheets1 & sheets2

        # Load once (performance fix)
        excel1 = pd.ExcelFile(file1)
        excel2 = pd.ExcelFile(file2)

        for sheet in common_sheets:
            try:
                df1 = excel1.parse(sheet).fillna("")
                df2 = excel2.parse(sheet).fillna("")

                diffs = compare_sheets(df1, df2, sheet)
                result["differences"].extend(diffs)

            except Exception as e:
                result["differences"].append({
                    "type": "Error",
                    "detail": f"{sheet}: {str(e)}"
                })

    except Exception as e:
        result["differences"].append({
            "type": "Error",
            "detail": f"File error: {str(e)}"
        })

    return result